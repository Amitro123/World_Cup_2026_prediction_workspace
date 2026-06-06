"""
תבנית אקסל — generate the Excel mirror of the workspace.

    python build_excel.py   ->   WorldCup2026_Template.xlsx

The Excel file reproduces the core logic with native formulas:
- Odds -> implied probability (American-odds conversion).
- A simplified live win/draw/loss recompute from minute + score.
- An on-track flag for each prediction.
Sheets: Teams, Matches, Odds, MyPredictions, LiveStatus.
"""

import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

DATA = os.path.join(os.path.dirname(__file__), "data")
OUT = os.path.join(os.path.dirname(__file__), "WorldCup2026_Template.xlsx")


def _dump(ws, df):
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)


def main():
    teams = pd.read_csv(os.path.join(DATA, "teams.csv"))
    matches = pd.read_csv(os.path.join(DATA, "matches.csv"))
    odds = pd.read_csv(os.path.join(DATA, "odds.csv"))
    preds = pd.read_csv(os.path.join(DATA, "my_predictions.csv"))

    wb = Workbook()
    ws_t = wb.active
    ws_t.title = "Teams"
    _dump(ws_t, teams[["team_id", "name_he", "name_en", "group_id",
                       "group_winner_odds", "tier", "power_rating"]])

    ws_m = wb.create_sheet("Matches")
    _dump(ws_m, matches[["match_id", "group_id", "home_id", "away_id",
                         "status", "minute", "home_goals", "away_goals",
                         "doc_pred_home", "doc_pred_away"]])

    # Odds sheet: implied probability from American odds, demonstrated for first
    # 4 teams as a reference (column F formula), rest carried as values.
    ws_o = wb.create_sheet("Odds")
    ws_o.append(["match_id", "p_home", "p_draw", "p_away"])
    for _, r in odds.iterrows():
        ws_o.append([r.match_id, r.p_home, r.p_draw, r.p_away])

    # Reference: American-odds -> implied probability on the Teams sheet
    ws_t.cell(row=1, column=9, value="implied_prob (formula)")
    for i in range(2, min(6, teams.shape[0] + 2)):  # first 4 rows as examples
        ws_t.cell(
            row=i, column=9,
            value=f'=IF(E{i}<0,(-E{i})/((-E{i})+100),100/(E{i}+100))',
        )

    # LiveStatus: live recompute mirror for 4 example matches.
    ws_l = wb.create_sheet("LiveStatus")
    ws_l.append([
        "match_id", "pick(H/D/A)", "minute", "home_goals", "away_goals",
        "p_home_live", "p_draw_live", "p_away_live",
        "prob_my_pick", "on_track?",
    ])
    # Simplified live mirror: blend pre-match odds with the realised result,
    # weighted by elapsed time. Explainable approximation of the Poisson engine.
    examples = ["A1", "H1", "I1", "L1"]
    for r_i, mid in enumerate(examples, start=2):
        o = odds.loc[odds.match_id == mid].iloc[0]
        p = preds.loc[preds.match_id == mid].iloc[0]
        m = matches.loc[matches.match_id == mid].iloc[0]
        ws_l.append([
            mid, p.pick, int(m.minute), 0, 0,
            None, None, None, None, None,
        ])
        rr = r_i
        # elapsed weight: w = minute/90
        # live prob = (1-w)*pre_match_prob + w*[1 if result matches else 0]
        ws_l.cell(row=rr, column=6, value=(  # p_home_live
            f'=(1-C{rr}/90)*{o.p_home}+(C{rr}/90)*IF(D{rr}>E{rr},1,0)'))
        ws_l.cell(row=rr, column=7, value=(  # p_draw_live
            f'=(1-C{rr}/90)*{o.p_draw}+(C{rr}/90)*IF(D{rr}=E{rr},1,0)'))
        ws_l.cell(row=rr, column=8, value=(  # p_away_live
            f'=(1-C{rr}/90)*{o.p_away}+(C{rr}/90)*IF(D{rr}<E{rr},1,0)'))
        ws_l.cell(row=rr, column=9, value=(  # prob of my pick
            f'=IF(B{rr}="H",F{rr},IF(B{rr}="D",G{rr},H{rr}))'))
        ws_l.cell(row=rr, column=10, value=(  # on track flag
            f'=IF(I{rr}>=0.55,"ON_TRACK",IF(I{rr}>=0.25,"AT_RISK","ALMOST_DEAD"))'))

    ws_p = wb.create_sheet("MyPredictions")
    _dump(ws_p, preds)

    wb.save(OUT)
    print(f"Saved {OUT}")
    print("Note: the LiveStatus sheet uses a time-weighted approximation of the")
    print("Poisson engine — change minute / goals and watch the flags update.")


if __name__ == "__main__":
    main()
